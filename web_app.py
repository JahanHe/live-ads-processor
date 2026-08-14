#!/usr/bin/env python3
import cgi
import json
import mimetypes
import os
import re
import shutil
import socket
import subprocess
import sys
import tempfile
import threading
import uuid
import webbrowser
import zipfile
from http import HTTPStatus
from http.server import BaseHTTPRequestHandler, ThreadingHTTPServer
from pathlib import Path
from urllib.parse import quote, unquote, urlparse

from openpyxl import load_workbook
from PIL import Image

from process_live_ads import (
    build_workbook,
    default_output_stem,
    read_csv_files,
    read_long_plan_files,
    rows_from_long_plan_values,
    to_number,
    validate_inputs,
    validate_long_plan_inputs,
)


def app_base_dir():
    if getattr(sys, "frozen", False):
        return Path(getattr(sys, "_MEIPASS", Path(sys.executable).parent))
    return Path(__file__).resolve().parent


APP_DIR = app_base_dir()
OUTPUT_DIR = Path(tempfile.gettempdir()) / "live_ads_panel_outputs"
STATIC_DIR = APP_DIR / "static"
TABLE_SUFFIXES = {".csv", ".tsv", ".txt", ".xlsx", ".xlsm", ".xls"}
OCR_DIR = Path(os.environ.get("LIVE_ADS_OCR_DIR", Path.home() / ".live_ads_processor" / "ocr"))


def summarize_rows(rows):
    spend_yuan = sum(to_number(row.get("总消耗")) for row in rows)
    totals = {
        "rows": len(rows),
        "spend_beans": spend_yuan * 10,
        "spend_yuan": spend_yuan,
        "deal_amount": sum(to_number(row.get("总成交金额")) for row in rows),
        "order_amount": sum(to_number(row.get("总下单金额")) for row in rows),
        "deal_orders": sum(to_number(row.get("总成交订单数")) for row in rows),
    }
    totals["deal_roi"] = totals["deal_amount"] / totals["spend_yuan"] if totals["spend_yuan"] else 0
    return totals


def form_bool(form, name, default=True):
    if name not in form:
        return default
    value = str(form.getfirst(name, "")).lower()
    return value in {"1", "true", "yes", "on"}


def safe_download_name(name):
    stem = Path(name).stem or "投放数据处理结果"
    stem = "".join(ch for ch in stem if ch not in '\\/:*?"<>|').strip()
    return f"{stem}_处理结果.xlsx"


def safe_xlsx_name(name):
    stem = Path(name).stem or "投放数据处理结果"
    stem = "".join(ch for ch in stem if ch not in '\\/:*?"<>|').strip()
    return f"{stem}.xlsx"


def form_files(form, name):
    if name not in form:
        return []
    value = form[name]
    files = value if isinstance(value, list) else [value]
    return [item for item in files if getattr(item, "filename", "")]


def save_uploaded_table(upload, prefix):
    suffix = Path(upload.filename).suffix.lower()
    if suffix not in TABLE_SUFFIXES:
        raise ValueError(f"{Path(upload.filename).name} 不是支持的表格文件。")
    temp_path = Path(tempfile.gettempdir()) / f"{prefix}{suffix}"
    with temp_path.open("wb") as f:
        f.write(upload.file.read())
    return temp_path


def save_uploaded_image(upload, prefix):
    temp_raw = Path(tempfile.gettempdir()) / f"{prefix}.upload"
    temp_png = Path(tempfile.gettempdir()) / f"{prefix}.png"
    with temp_raw.open("wb") as f:
        f.write(upload.file.read())
    try:
        Image.open(temp_raw).convert("RGB").save(temp_png)
    except Exception as exc:
        raise ValueError(f"{Path(upload.filename).name} 不是可识别的图片文件。") from exc
    finally:
        try:
            temp_raw.unlink()
        except FileNotFoundError:
            pass
    return temp_png


def sheet_preview(workbook_path, sheet_name):
    wb = load_workbook(workbook_path, data_only=True)
    ws = wb[sheet_name]
    rows = []
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column, values_only=True):
        rows.append(["" if value is None else value for value in row])
    return rows


def ocr_number_after(text, labels):
    for label in labels:
        pattern = rf"{re.escape(label)}[\s:：¥￥]*([0-9][0-9,]*(?:\.[0-9]+)?)"
        match = re.search(pattern, text)
        if match:
            return match.group(1)
    return ""


def ocr_numbers(line):
    return re.findall(r"[¥￥]?\s*([0-9][0-9,]*(?:\.[0-9]+)?)", line)


def ocr_next_numbers(lines, required_labels):
    for index, line in enumerate(lines[:-1]):
        if all(label in line for label in required_labels):
            return ocr_numbers(lines[index + 1])
    return []


def parse_long_plan_ocr_text(text):
    lines = [line.strip() for line in text.splitlines() if line.strip()]
    result = {
        "消耗总金额": ocr_number_after(text, ["消耗总金额", "直播间消耗", "直播间消耗耗"]),
        "曝光总人数": ocr_number_after(text, ["曝光总人数", "直播间曝光人数"]),
        "进入总人数": ocr_number_after(text, ["进入总人数", "直播间观看人数"]),
        "点赞总次数": ocr_number_after(text, ["点赞总次数", "直播间点赞次数"]),
        "评论总次数": ocr_number_after(text, ["评论总次数", "直播间评论次数"]),
        "新增总关注": ocr_number_after(text, ["新增总关注", "新增总关注数", "新增总粉丝"]),
        "成交GMV": ocr_number_after(text, ["当场成交GMV", "直接成交GMV", "净成交金额", "成交GMV"]),
        "成交订单数": ocr_number_after(text, ["当场成交订单数", "直接成交订单数", "净成交订单数", "成交订单数"]),
        "下单GMV": ocr_number_after(text, ["当场下单GMV", "直接下单GMV", "下单GMV"]),
        "下单订单数": ocr_number_after(text, ["当场下单订单数", "直接下单订单数", "下单订单数"]),
    }
    live_numbers = ocr_next_numbers(lines, ["消耗总金额", "曝光总人数", "新增总关注"])
    if len(live_numbers) >= 6:
        result.update(
            {
                "消耗总金额": live_numbers[0],
                "曝光总人数": live_numbers[1],
                "进入总人数": live_numbers[2],
                "点赞总次数": live_numbers[3],
                "评论总次数": live_numbers[4],
                "新增总关注": live_numbers[5],
            }
        )
    deal_numbers = ocr_next_numbers(lines, ["成交ROI", "当场成交GMV", "当场成交订单数"])
    if len(deal_numbers) >= 6:
        result["成交GMV"] = deal_numbers[2]
        result["成交订单数"] = deal_numbers[4]
    order_numbers = ocr_next_numbers(lines, ["净成交订单数", "当场下单GMV", "当场下单订单数"])
    if len(order_numbers) >= 6:
        result["下单GMV"] = order_numbers[3]
        result["下单订单数"] = order_numbers[5]
    return result


def tesseract_path():
    installed = OCR_DIR / ("tesseract.exe" if sys.platform == "win32" else "tesseract")
    if installed.exists():
        return str(installed)
    bundled = APP_DIR / "ocr" / ("tesseract.exe" if sys.platform == "win32" else "tesseract")
    if bundled.exists():
        return str(bundled)
    system = shutil.which("tesseract")
    if system:
        return system
    if sys.platform == "win32":
        for path in (
            Path(os.environ.get("LOCALAPPDATA", "")) / "Programs" / "Tesseract-OCR" / "tesseract.exe",
            Path("C:/Program Files/Tesseract-OCR/tesseract.exe"),
            Path("C:/Program Files (x86)/Tesseract-OCR/tesseract.exe"),
        ):
            if path.exists():
                return str(path)
    return None


def tessdata_dir():
    for path in (OCR_DIR / "tessdata", APP_DIR / "ocr" / "tessdata"):
        if (path / "chi_sim.traineddata").exists() and (path / "eng.traineddata").exists():
            return path
    tesseract = tesseract_path()
    if tesseract:
        share = Path(tesseract).resolve().parents
        for parent in share:
            candidate = parent / "share" / "tessdata"
            if (candidate / "chi_sim.traineddata").exists() and (candidate / "eng.traineddata").exists():
                return candidate
    for path in (Path("/opt/homebrew/share/tessdata"), Path("/usr/local/share/tessdata")):
        if (path / "chi_sim.traineddata").exists() and (path / "eng.traineddata").exists():
            return path
    if sys.platform == "win32":
        for path in (
            Path(os.environ.get("LOCALAPPDATA", "")) / "Programs" / "Tesseract-OCR" / "tessdata",
            Path("C:/Program Files/Tesseract-OCR/tessdata"),
            Path("C:/Program Files (x86)/Tesseract-OCR/tessdata"),
        ):
            if (path / "chi_sim.traineddata").exists() and (path / "eng.traineddata").exists():
                return path
    return None


def tesseract_env():
    env = os.environ.copy()
    tessdata = tessdata_dir()
    if tessdata:
        env["TESSDATA_PREFIX"] = str(tessdata)
    return env


def ocr_status_payload():
    tesseract = tesseract_path()
    tessdata = tessdata_dir()
    installed = bool(tesseract and tessdata)
    return {
        "ok": True,
        "installed": installed,
        "tesseractPath": str(tesseract or ""),
        "tessdataPath": str(tessdata or ""),
        "installable": not installed and ocr_install_available(),
        "installLabel": "" if installed else ocr_install_label(),
    }


def ocr_component_packages():
    suffix = "windows" if sys.platform == "win32" else "mac" if sys.platform == "darwin" else "linux"
    return [
        APP_DIR / f"ocr-component-{suffix}.zip",
        APP_DIR / "ocr-component.zip",
    ]


def ocr_install_available():
    if any(path.exists() for path in ocr_component_packages()):
        return True
    if sys.platform == "darwin":
        return shutil.which("brew") is not None
    if sys.platform == "win32":
        return shutil.which("winget") is not None
    return False


def ocr_install_label():
    if any(path.exists() for path in ocr_component_packages()):
        return "安装截图识别组件"
    if sys.platform == "darwin" and shutil.which("brew"):
        return "安装截图识别组件"
    if sys.platform == "win32" and shutil.which("winget"):
        return "安装截图识别组件"
    return ""


def install_ocr_component():
    if tesseract_path() and tessdata_dir():
        return {"ok": True, "installed": True, "message": "截图识别组件已安装。"}
    for package in ocr_component_packages():
        if package.exists():
            OCR_DIR.mkdir(parents=True, exist_ok=True)
            with zipfile.ZipFile(package) as archive:
                archive.extractall(OCR_DIR)
            executable = OCR_DIR / ("tesseract.exe" if sys.platform == "win32" else "tesseract")
            if executable.exists() and sys.platform != "win32":
                executable.chmod(executable.stat().st_mode | 0o111)
            if tesseract_path() and tessdata_dir():
                return {"ok": True, "installed": True, "message": "截图识别组件已安装。"}
            raise RuntimeError("识别组件包已解压，但缺少程序或中文识别数据。")
    if sys.platform == "darwin":
        brew = shutil.which("brew")
        if not brew:
            raise RuntimeError("当前电脑没有可用的组件包，也没有 Homebrew，无法一键安装。")
        subprocess.run([brew, "install", "tesseract", "tesseract-lang"], check=True, text=True, capture_output=True, timeout=600)
        if not (tesseract_path() and tessdata_dir()):
            raise RuntimeError("组件安装完成，但当前进程还没有检测到截图识别组件。请重启软件后再试。")
        return {"ok": True, "installed": True, "message": "截图识别组件已安装。"}
    if sys.platform == "win32":
        winget = shutil.which("winget")
        if not winget:
            raise RuntimeError("当前电脑没有可用的组件包，也没有 winget，无法一键安装。")
        subprocess.run(
            [winget, "install", "--id", "UB-Mannheim.TesseractOCR", "-e", "--accept-package-agreements", "--accept-source-agreements"],
            check=True,
            text=True,
            capture_output=True,
            timeout=600,
        )
        if not (tesseract_path() and tessdata_dir()):
            raise RuntimeError("组件安装完成，但当前进程还没有检测到截图识别组件。请重启软件后再试。")
        return {"ok": True, "installed": True, "message": "截图识别组件已安装。"}
    raise RuntimeError("当前系统暂不支持一键安装截图识别组件。")


def missing_tesseract_payload():
    if sys.platform == "darwin":
        command = "brew install tesseract tesseract-lang"
        hint = "当前 Mac 还不能自动识别截图。你可以先手动填写；也可以一键安装识别组件后继续使用。"
    elif sys.platform == "win32":
        command = "把 tesseract.exe 放到 ocr/tesseract.exe，或安装 Tesseract 后加入 PATH"
        hint = "当前 Windows 还不能自动识别截图。你可以先手动填写；正式桌面版可以把识别组件一起打包。"
    else:
        command = "安装 tesseract-ocr 和中文语言包，例如 apt install tesseract-ocr tesseract-ocr-chi-sim"
        hint = "当前电脑还不能自动识别截图。你可以先手动填写；安装识别组件和中文语言包后重启面板即可。"
    return {
        "ok": False,
        "code": "missing_tesseract",
        "error": "截图识别组件未安装。",
        "hint": hint,
        "installCommand": command,
        "canInstall": ocr_install_available(),
        "installLabel": ocr_install_label(),
    }


class AppHandler(BaseHTTPRequestHandler):
    server_version = "LiveAdsPanel/1.0"

    def do_GET(self):
        parsed = urlparse(self.path)
        if parsed.path == "/":
            self.serve_file(STATIC_DIR / "index.html", "text/html; charset=utf-8")
            return
        if parsed.path == "/api/ocr-status":
            self.send_json(ocr_status_payload())
            return
        if parsed.path.startswith("/static/"):
            rel_path = unquote(parsed.path.removeprefix("/static/"))
            self.serve_static(rel_path)
            return
        if parsed.path.startswith("/download/"):
            self.serve_download(parsed.path.removeprefix("/download/"))
            return
        if parsed.path.startswith("/image/"):
            self.serve_image(parsed.path.removeprefix("/image/"))
            return
        self.send_error(HTTPStatus.NOT_FOUND, "Not Found")

    def do_POST(self):
        path = urlparse(self.path).path
        if path == "/api/ocr-long-plan":
            self.handle_ocr_long_plan()
            return
        if path == "/api/install-ocr":
            self.handle_install_ocr()
            return
        if path != "/api/process":
            self.send_error(HTTPStatus.NOT_FOUND, "接口不存在")
            return

        content_type = self.headers.get("content-type", "")
        if "multipart/form-data" not in content_type:
            self.send_json({"ok": False, "error": "请上传表格文件。"}, HTTPStatus.BAD_REQUEST)
            return

        form = cgi.FieldStorage(
            fp=self.rfile,
            headers=self.headers,
            environ={
                "REQUEST_METHOD": "POST",
                "CONTENT_TYPE": content_type,
                "CONTENT_LENGTH": self.headers.get("content-length", "0"),
            },
        )
        uploads = form_files(form, "files") or form_files(form, "file")
        if not uploads:
            self.send_json({"ok": False, "error": "没有收到文件。"}, HTTPStatus.BAD_REQUEST)
            return

        job_id = uuid.uuid4().hex
        job_dir = OUTPUT_DIR / job_id
        job_dir.mkdir(parents=True, exist_ok=True)
        temp_csvs = []
        temp_long_plans = []
        temp_images = []
        custom_output_name = str(form.getfirst("outputName", "")).strip()
        use_thousands = form_bool(form, "useThousands", True)
        transpose_summary = form_bool(form, "transposeSummary", True)
        remove_zero_columns = form_bool(form, "removeZeroColumns", False)
        rate_metrics_as_percent = form_bool(form, "rateMetricsAsPercent", True)
        decimal_mode = str(form.getfirst("decimalMode", "fixed2"))
        if decimal_mode not in {"fixed2", "full"}:
            decimal_mode = "fixed2"

        try:
            for idx, upload in enumerate(uploads, 1):
                temp_csvs.append(save_uploaded_table(upload, f"live_ads_{job_id}_{idx}"))
            long_plans = form_files(form, "longPlans")
            for idx, plan in enumerate(long_plans, 1):
                temp_long_plans.append(save_uploaded_table(plan, f"live_ads_{job_id}_long_plan_{idx}"))
            screenshots = form_files(form, "screenshots")
            for idx, screenshot in enumerate(screenshots, 1):
                temp_images.append(save_uploaded_image(screenshot, f"live_ads_{job_id}_screenshot_{idx}"))

            header_sets, rows = read_csv_files(temp_csvs)
            if not rows:
                raise ValueError("表格没有数据行。")
            warnings = validate_inputs(header_sets, rows)
            if temp_long_plans:
                long_header_sets, long_rows = read_long_plan_files(temp_long_plans)
                warnings.extend(validate_long_plan_inputs(long_header_sets))
                rows.extend(long_rows)
            manual_long_rows = rows_from_long_plan_values(json.loads(form.getfirst("longPlanRow", "{}")))
            rows.extend(manual_long_rows)
            output_name = safe_xlsx_name(custom_output_name or default_output_stem(rows))
            output_path = job_dir / output_name
            combined_image_name = f"{Path(output_name).stem}_拼接图.png"
            combined_image_path = job_dir / combined_image_name
            result = build_workbook(
                temp_csvs,
                output_path,
                use_thousands=use_thousands,
                decimal_mode=decimal_mode,
                transpose_summary=transpose_summary,
                remove_zero_columns=remove_zero_columns,
                image_paths=temp_images,
                combined_image_path=combined_image_path,
                rate_metrics_as_percent=rate_metrics_as_percent,
                long_plan_paths=temp_long_plans,
                long_plan_rows=manual_long_rows,
            )
            warnings = warnings or result.get("warnings", [])
            summary = summarize_rows(rows)
            previews = {
                "sheet2": sheet_preview(output_path, "数据汇总"),
                "sheet3": sheet_preview(output_path, "结算整理表"),
            }
        except Exception as exc:
            self.send_json({"ok": False, "error": str(exc)}, HTTPStatus.BAD_REQUEST)
            return
        finally:
            try:
                for temp_path in temp_csvs + temp_long_plans + temp_images:
                    temp_path.unlink()
            except FileNotFoundError:
                pass

        self.send_json(
            {
                "ok": True,
                "fileName": output_name,
                "downloadUrl": f"/download/{job_id}/{quote(output_name)}",
                "summary": summary,
                "previews": previews,
                "warnings": warnings,
                "combinedImageUrl": f"/image/{job_id}/{quote(combined_image_name)}" if result.get("combined_image") else "",
                "combinedImageName": combined_image_name if result.get("combined_image") else "",
            }
        )

    def handle_ocr_long_plan(self):
        tesseract = tesseract_path()
        if not tesseract:
            self.send_json(missing_tesseract_payload(), HTTPStatus.BAD_REQUEST)
            return
        content_type = self.headers.get("content-type", "")
        if "multipart/form-data" not in content_type:
            self.send_json({"ok": False, "error": "请上传长期计划截图。"}, HTTPStatus.BAD_REQUEST)
            return
        form = cgi.FieldStorage(
            fp=self.rfile,
            headers=self.headers,
            environ={
                "REQUEST_METHOD": "POST",
                "CONTENT_TYPE": content_type,
                "CONTENT_LENGTH": self.headers.get("content-length", "0"),
            },
        )
        images = form_files(form, "ocrImage")
        if not images:
            self.send_json({"ok": False, "error": "没有收到截图。"}, HTTPStatus.BAD_REQUEST)
            return
        image = images[0]
        temp_image = None
        try:
            temp_image = save_uploaded_image(image, f"live_ads_ocr_{uuid.uuid4().hex}")
            result = subprocess.run(
                [tesseract, str(temp_image), "stdout", "-l", "chi_sim+eng", "--psm", "6"],
                check=False,
                capture_output=True,
                env=tesseract_env(),
                text=True,
                timeout=20,
            )
            if result.returncode != 0:
                raise RuntimeError(result.stderr.strip() or "OCR 识别失败。")
            self.send_json({"ok": True, "fields": parse_long_plan_ocr_text(result.stdout), "text": result.stdout})
        except Exception as exc:
            self.send_json({"ok": False, "error": str(exc)}, HTTPStatus.BAD_REQUEST)
        finally:
            try:
                if temp_image:
                    temp_image.unlink()
            except FileNotFoundError:
                pass

    def handle_install_ocr(self):
        try:
            payload = install_ocr_component()
            payload.update(ocr_status_payload())
            payload["message"] = payload.get("message") or "截图识别组件已安装。"
            self.send_json(payload)
        except Exception as exc:
            self.send_json(
                {
                    "ok": False,
                    "code": "ocr_install_failed",
                    "error": str(exc),
                    "status": ocr_status_payload(),
                },
                HTTPStatus.BAD_REQUEST,
            )

    def serve_file(self, path, content_type=None):
        if not path.exists() or not path.is_file():
            self.send_error(HTTPStatus.NOT_FOUND, "File Not Found")
            return
        data = path.read_bytes()
        content_type = content_type or mimetypes.guess_type(path.name)[0] or "application/octet-stream"
        self.send_response(HTTPStatus.OK)
        self.send_header("Content-Type", content_type)
        self.send_header("Content-Length", str(len(data)))
        self.end_headers()
        self.wfile.write(data)

    def serve_static(self, rel_path):
        full_path = (STATIC_DIR / rel_path).resolve()
        if STATIC_DIR.resolve() not in full_path.parents and full_path != STATIC_DIR.resolve():
            self.send_error(HTTPStatus.FORBIDDEN, "Forbidden")
            return
        self.serve_file(full_path)

    def serve_download(self, rel_path):
        full_path = (OUTPUT_DIR / unquote(rel_path)).resolve()
        if OUTPUT_DIR.resolve() not in full_path.parents:
            self.send_error(HTTPStatus.FORBIDDEN, "Forbidden")
            return
        if not full_path.exists():
            self.send_error(HTTPStatus.NOT_FOUND, "File Not Found")
            return
        data = full_path.read_bytes()
        encoded_name = quote(full_path.name)
        self.send_response(HTTPStatus.OK)
        self.send_header(
            "Content-Type",
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        self.send_header("Content-Length", str(len(data)))
        self.send_header(
            "Content-Disposition",
            f"attachment; filename=\"output.xlsx\"; filename*=UTF-8''{encoded_name}",
        )
        self.end_headers()
        self.wfile.write(data)

    def serve_image(self, rel_path):
        full_path = (OUTPUT_DIR / unquote(rel_path)).resolve()
        if OUTPUT_DIR.resolve() not in full_path.parents:
            self.send_error(HTTPStatus.FORBIDDEN, "Forbidden")
            return
        if not full_path.exists():
            self.send_error(HTTPStatus.NOT_FOUND, "File Not Found")
            return
        data = full_path.read_bytes()
        self.send_response(HTTPStatus.OK)
        self.send_header("Content-Type", mimetypes.guess_type(full_path.name)[0] or "image/png")
        self.send_header("Content-Length", str(len(data)))
        self.send_header("Content-Disposition", f"attachment; filename=\"summary.png\"; filename*=UTF-8''{quote(full_path.name)}")
        self.end_headers()
        self.wfile.write(data)

    def send_json(self, payload, status=HTTPStatus.OK):
        data = json.dumps(payload, ensure_ascii=False).encode("utf-8")
        self.send_response(status)
        self.send_header("Content-Type", "application/json; charset=utf-8")
        self.send_header("Content-Length", str(len(data)))
        self.end_headers()
        self.wfile.write(data)

    def log_message(self, fmt, *args):
        print(f"{self.address_string()} - {fmt % args}")


def main():
    port = find_available_port(int(os.environ.get("PORT", "8765")))
    server = ThreadingHTTPServer(("127.0.0.1", port), AppHandler)
    url = f"http://127.0.0.1:{port}"
    if os.environ.get("NO_BROWSER") != "1":
        threading.Timer(0.8, lambda: webbrowser.open(url)).start()
    print(f"投放数据处理面板已启动：{url}")
    server.serve_forever()


def find_available_port(start_port):
    for port in range(start_port, start_port + 50):
        with socket.socket(socket.AF_INET, socket.SOCK_STREAM) as sock:
            sock.setsockopt(socket.SOL_SOCKET, socket.SO_REUSEADDR, 1)
            try:
                sock.bind(("127.0.0.1", port))
            except OSError:
                continue
            return port
    raise RuntimeError("没有找到可用端口。")


if __name__ == "__main__":
    main()
