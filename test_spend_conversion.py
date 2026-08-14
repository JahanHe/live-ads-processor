import csv
import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook, load_workbook

from process_live_ads import SOURCE_HEADERS, aggregate_rows, build_workbook, populate_summary, summary_values
from web_app import parse_long_plan_ocr_text, summarize_rows


class SpendConversionTest(unittest.TestCase):
    def setUp(self):
        self.rows = [{"名称": "roi", "总消耗": "415.2", "总成交金额": "830.4"}]

    def test_input_spend_is_yuan(self):
        values = summary_values(aggregate_rows(self.rows), "")

        self.assertEqual(values["总消耗金额"], 415.2)
        self.assertEqual(values["总消耗（豆）"], 4152)
        self.assertEqual(values["总成交roi"], 2)
        self.assertEqual(summarize_rows(self.rows)["spend_yuan"], 415.2)

    def test_excel_formulas_do_not_divide_input_spend(self):
        vertical = Workbook().active
        populate_summary(vertical, 2, transpose=True)
        self.assertEqual(vertical["B2"].value, "=B3*10")
        self.assertEqual(vertical["B3"].value, "=SUM('直播投放数据源'!$H$2:$H$2)")

        horizontal = Workbook().active
        populate_summary(horizontal, 2, transpose=False)
        self.assertEqual(horizontal["B2"].value, "=C2*10")
        self.assertEqual(horizontal["C2"].value, "=SUM('直播投放数据源'!$H$2:$H$2)")

    def test_long_plan_rows_merge_with_standard_csv(self):
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            standard = root / "standard.csv"
            plan = root / "plan.csv"
            output = root / "output.xlsx"
            with standard.open("w", encoding="utf-8-sig", newline="") as f:
                writer = csv.DictWriter(f, fieldnames=SOURCE_HEADERS)
                writer.writeheader()
                writer.writerow({"名称": "标准订单", "总消耗": 100, "总成交金额": 10})
            with plan.open("w", encoding="utf-8-sig", newline="") as f:
                writer = csv.DictWriter(f, fieldnames=["消耗总金额", "曝光总人数", "当场成交GMV"])
                writer.writeheader()
                writer.writerow({"消耗总金额": "¥1416.5", "曝光总人数": 4700, "当场成交GMV": "¥5553.55"})

            build_workbook([standard], output, long_plan_paths=[plan])

            wb = load_workbook(output, data_only=True)
            summary = wb["数据汇总"]
            values = {summary.cell(row, 1).value: summary.cell(row, 2).value for row in range(2, summary.max_row + 1)}
            self.assertEqual(wb["直播投放数据源"].max_row, 3)
            self.assertEqual(values["总消耗金额"], 1516.5)
            self.assertEqual(values["总消耗（豆）"], 15165)
            self.assertEqual(values["曝光人数"], 4700)
            self.assertEqual(values["总成交金额"], 5563.55)

    def test_long_plan_ocr_handles_noisy_full_page_capture(self):
        text = """
直播间加热效果 —- 2026-03-05 = 2026-08-13 f=}
消耗总金额               BSA               进入总人数               点跑总次数               评论总次数               新增总关注
¥81293.34             312595              73075                7476                4062                1005
电商加热效果 —«- 2026-03-05 = 2026-08-13 [=]
成交ROI               直接成交ROI              当场成交GMV             直接成交GMV             当场成交订单数            直接成交订单数
6.4698              6.4698            ¥627835.47          ¥627835.47             3369                3369
净成交订单数              净成交金额               净成交ROI               当场下单GMV             直接下单GMV             当场下单订单数
2837             ¥466428.57            4.8065            ¥746545.58          ¥746545.58             3725
直接下单订单数              商品点击人数              商品点击次数
3725                6843                17027
"""

        fields = parse_long_plan_ocr_text(text)

        self.assertEqual(fields["消耗总金额"], "81293.34")
        self.assertEqual(fields["曝光总人数"], "312595")
        self.assertEqual(fields["进入总人数"], "73075")
        self.assertEqual(fields["点赞总次数"], "7476")
        self.assertEqual(fields["评论总次数"], "4062")
        self.assertEqual(fields["新增总关注"], "1005")
        self.assertEqual(fields["成交GMV"], "627835.47")
        self.assertEqual(fields["成交订单数"], "3369")
        self.assertEqual(fields["下单GMV"], "746545.58")
        self.assertEqual(fields["下单订单数"], "3725")

    def test_long_plan_ocr_handles_cropped_live_block(self):
        text = """
直播间加热效果 —«-2026-08-12 = 2026-08-12 =)
消耗总金额                 HBA                 进入总人数                 点跑总次数                 评论总次数                 新增总关注
¥729.96                 6322                  1500                   360                    39                     17
电商加热效果 —«- 2026-03-05 = 2026-08-13 [=]
成交ROI                 直接成交ROI               当场成交GMV               直接成交GMV              当场成交订单数              直接成交订单数
6.4698                6.4698             ¥627835.47           ¥627835.47              3369                  3369
净成交订单数                净成交金额                 净成交ROI                当场下单GMV               直接下单GMV              当场下单订单数
2837              ¥466428.57             4.8065             ¥746545.58           ¥746545.58               3725
直接下单订单数               商品点击人数               商品点击次数
3725                  6843                 17027
"""

        fields = parse_long_plan_ocr_text(text)

        self.assertEqual(fields["消耗总金额"], "729.96")
        self.assertEqual(fields["曝光总人数"], "6322")
        self.assertEqual(fields["进入总人数"], "1500")
        self.assertEqual(fields["点赞总次数"], "360")
        self.assertEqual(fields["评论总次数"], "39")
        self.assertEqual(fields["新增总关注"], "17")
        self.assertEqual(fields["成交GMV"], "627835.47")
        self.assertEqual(fields["成交订单数"], "3369")
        self.assertEqual(fields["下单GMV"], "746545.58")
        self.assertEqual(fields["下单订单数"], "3725")


if __name__ == "__main__":
    unittest.main()
